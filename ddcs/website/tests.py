from io import BytesIO

from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from ddcs.website.context_processors import analytics
from ddcs.website.dashboard_export import (
    build_dashboard_export_meta,
    nationwide_export_meta,
)
from ddcs.website.methods_lists import (
    get_methods_lists_payload,
    load_monitored_accounts,
    load_monitored_keywords,
)


class WebsiteTests(TestCase):
    def setUp(self):
        self.client = Client()

    def test_dfdw_2025_page(self):
        """Test that the DFDW 2025 page returns a 200 status code."""
        url = reverse("website:dfdw_2025")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    @override_settings(DEBUG=True)
    def test_public_dashboard_party_videos_carousel(self):
        response = self.client.get(reverse("website:public_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="partyVideosCarousel"')
        self.assertContains(response, "Videos über die Zeit")
        self.assertContains(response, "Videos gesamt")
        self.assertIsNotNone(
            response.context["temporal_party_distribution_plot"]["html"]
        )
        self.assertIsNotNone(response.context["party_distribution_plot"]["html"])
        self.assertContains(
            response, 'data-export-title="Videos von Partei-Accounts nach Partei"'
        )
        self.assertContains(response, "ganz Deutschland")
        self.assertContains(response, "1. Juli 2026")


class DashboardExportMetaTests(TestCase):
    def test_nationwide_scope_and_period(self):
        meta = build_dashboard_export_meta(
            video_stats={
                "start_date": "2026-07-01",
                "end_date": "2026-08-14",
                "n_accounts": 127,
            }
        )
        caption = meta["videos_gesamt"]["caption"]
        self.assertEqual(
            meta["videos_gesamt"]["title"],
            "Videos von Partei-Accounts nach Partei",
        )
        self.assertIn("1. Juli 2026", caption)
        self.assertIn("14. August 2026", caption)
        self.assertIn("127 Accounts", caption)
        self.assertIn("ganz Deutschland", caption)
        self.assertIn("Bundes- und Landesverbände", caption)
        self.assertIn("Kandidierende", caption)
        self.assertIn("Grafik: Dein Feed, Deine Wahl.", caption)
        self.assertIn("\nQuelle: TikTok Research API.\n", caption)
        self.assertNotIn("\n\n", caption)

    def test_berlin_scope_uses_state_accounts(self):
        caption = build_dashboard_export_meta(
            video_stats={"start_date": "2026-07-01", "end_date": "2026-08-14"},
            selected_bundesland="BE",
        )["likes_gesamt"]["caption"]
        self.assertIn("Berlin zugeordneten Partei-Accounts", caption)
        self.assertIn("Abgeordnetenhaus von Berlin", caption)
        self.assertNotIn("ganz Deutschland", caption)


class NationwideExportMetaTests(TestCase):
    def test_includes_homepage_plot_keys(self):
        meta = nationwide_export_meta()
        self.assertIn("videos_gesamt", meta)
        self.assertIn("videos_zeit", meta)
        caption = meta["videos_gesamt"]["caption"]
        self.assertIn("Grafik: Dein Feed, Deine Wahl.", caption)


class MethodsListsTests(TestCase):
    def setUp(self):
        self.client = Client()

    def test_methods_lists_page(self):
        response = self.client.get(reverse("website:methods_lists"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Keywords")
        self.assertContains(response, "Accounts")
        self.assertGreater(response.context["keyword_count"], 0)
        self.assertGreater(response.context["account_count"], 0)

    def test_methods_lists_json_export(self):
        response = self.client.get(reverse("website:methods_lists_json"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        data = response.json()
        self.assertIn("keywords", data)
        self.assertIn("accounts", data)
        self.assertEqual(data["keywords"], load_monitored_keywords())
        self.assertEqual(
            data["accounts"][0].keys(),
            {"username", "party", "bundesland"},
        )

    def test_methods_lists_xlsx_export(self):
        response = self.client.get(reverse("website:methods_lists_xlsx"))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "spreadsheetml.sheet",
            response["Content-Type"],
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ["keywords", "accounts"])
        self.assertEqual(wb["keywords"]["A1"].value, "keyword")
        self.assertEqual(
            [c.value for c in wb["accounts"][1]],
            ["username", "party", "bundesland"],
        )

    def test_payload_matches_fixture_loaders(self):
        payload = get_methods_lists_payload()
        self.assertEqual(len(payload["keywords"]), len(load_monitored_keywords()))
        self.assertEqual(len(payload["accounts"]), len(load_monitored_accounts()))
        sample = next(a for a in payload["accounts"] if a["username"] == "insidecdu")
        self.assertEqual(sample["party"], "CDU/CSU")
        self.assertTrue(sample["bundesland"])
        minor = next(a for a in payload["accounts"] if a["username"] == "piratenpartei")
        self.assertEqual(minor["party"], "Sonstige")


class AnalyticsProcessorTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @override_settings(
        ANALYTICS_SCRIPT_SRC="https://example.com", ANALYTICS_WEBSITE_ID="12345"
    )
    def test_analytics_processor_returns_correct_values(self):
        request = self.factory.get("/")
        result = analytics(request)

        self.assertEqual(result["ANALYTICS_SCRIPT_SRC"], "https://example.com")
        self.assertEqual(result["ANALYTICS_WEBSITE_ID"], "12345")

    @override_settings(ANALYTICS_WEBSITE_ID=None)
    def test_analytics_processor_handles_none_values(self):
        request = self.factory.get("/")
        result = analytics(request)

        self.assertIsNone(result["ANALYTICS_WEBSITE_ID"])
